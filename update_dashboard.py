"""Refresh the website JSON from the portfolio analysis workbook."""

import json
import os
import sys
from datetime import date, datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is not installed. Run: py -m pip install openpyxl")
    sys.exit(1)


WORKBOOK = Path(r"D:\Pyhton code\FINAL Scripts\portfolio_analysis.xlsx")
OUTPUT = Path(__file__).resolve().parent / "public" / "stocks.json"


def json_value(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def main():
    if not WORKBOOK.exists():
        print(f"ERROR: Workbook not found: {WORKBOOK}")
        return 1

    print(f"Reading: {WORKBOOK}")
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    if "Summary" not in workbook.sheetnames:
        print("ERROR: The workbook does not contain a Summary sheet.")
        return 1

    sheet = workbook["Summary"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    records = []

    for row in rows:
        if not any(value is not None for value in row):
            continue
        records.append({
            headers[index]: json_value(value)
            for index, value in enumerate(row)
            if index < len(headers) and headers[index]
        })

    payload = {
        "updated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "source": str(WORKBOOK),
        "summary": records,
    }
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    temporary = OUTPUT.with_suffix(".json.tmp")
    temporary.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    os.replace(temporary, OUTPUT)
    print(f"Website data updated: {len(records)} securities")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
